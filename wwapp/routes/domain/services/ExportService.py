from io import BytesIO
from flask import session
from routes.domain.repositories.RouteRepository import RouteRepository
from routes.domain.repositories.TemplateRepository import TemplateRepository
from routes.usecases.ForExportingRoutes import ForExportingRoutes
from routes.usecases.ForReadingRoutes import ForReadingRoutes
from routes.usecases.ForSavingRoutes import ForSavingRoutes
import openpyxl

from users.usecases.ForReadingUsers import ForReadingUsers

class ExportService(ForExportingRoutes):
    def __init__(self, 
                 template_repository: TemplateRepository, 
                 route_repository: RouteRepository,
                 for_reading_users: ForReadingUsers):
        self.template_repository = template_repository
        self.route_repository = route_repository
        self.for_reading_users = for_reading_users

    def export(self, year: int) -> bytes: 
        template = self.template_repository.get_latest_template()
        fileXLSX = openpyxl.load_workbook(template.file_data)
        sheet = fileXLSX["Protokoll über die Betreuung"]
        routes = self.route_repository.get_all_routes_of_year(year)
        sheet.cell(row=2, column=4).value = year 
        i = 12
        for route in routes:
            #if (route.participants is None) or (len(route.participants) == 0):
             #   continue
            user = self.for_reading_users.get_user_by_id(route.user_id)
            if (route.persons is None) or (route.persons == ''):                
                user_names = user.first_name + " " + user.last_name
            else:
                user_names = route._persons
            sheet.cell(row=i, column=1).value = route.date.strftime('%d.%m.%Y')
            sheet.cell(row=i, column=2).value = route.workplace            
            sheet.cell(row=i, column=3).value = user_names
            sheet.cell(row=i, column=4).value = route.activity
            sheet.cell(row=i, column=5).value = route.description
            if (route.type == 'rotblau'):    
                sheet.cell(row=i, column=6).value = route.diamond_count
                sheet.cell(row=i, column=7).value = route.arrow_count
                sheet.cell(row=i, column=8).value = route.sticker_count
                sheet.cell(row=i, column=9).value = route.working_hours
            elif (route.type == 'gelb'):
                sheet.cell(row=i, column=10).value = route.diamond_count
                sheet.cell(row=i, column=11).value = route.arrow_count
                sheet.cell(row=i, column=12).value = route.sticker_count
                sheet.cell(row=i, column=13).value = route.working_hours

            i += 1
        
        io = BytesIO()
        fileXLSX.save(io)
        content = io.getvalue()
        return content
